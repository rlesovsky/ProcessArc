"""IO List Exporter (Plan §9) — render the Project Device Model into the UFP template.

Behavior:
  - Loads the supplied UFP IO-list template, preserving its sheet layout,
    metadata rows (UDT Type / Folder), and headers (Plan §7.3).
  - Clears any existing device rows below the metadata block.
  - For each non-excluded device in the Device Model, finds the matching
    template sheet (by DeviceClass → sheet name) and writes the identity
    columns (System Name, System Number, Name) — and Description if the
    template has such a column.
  - Register columns are LEFT BLANK (Plan §9A.5 safe default — Q11 gates
    standard register pre-fill on PLC-team confirmation).
  - Excluded devices (`review_status == EXCLUDED`) are not written.
  - A sheet with no devices for its class is left header-only.
  - Output saved as `{SiteName}_Ignition_IOList.xlsx`.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl.workbook import Workbook

from backend.model import DeviceClass, DeviceModel, ReviewStatus
from backend.model.device import DeviceRecord
from backend.model.plant import PlantConfiguration, TankRecord
from backend.naming import ignition_name
from backend.profiles.ufp_registers import TANK_REGISTER_PATTERN, tank_register_for

from .template import (
	ColumnKind,
	SHEET_FOR_CLASS,
	TemplateMap,
	TemplateSheet,
	load_template_workbook,
	read_template,
)


def _safe_site_name(plant: PlantConfiguration) -> str:
	"""File-system-safe variant of the site name for the export filename."""
	base = plant.site_name.strip() or "Project"
	return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in base)


def _devices_for_sheet(model: DeviceModel, sheet: TemplateSheet) -> list[DeviceRecord]:
	# Map the sheet name back to the device class. Match is case-insensitive
	# on the canonical UFP name.
	want = sheet.name.lower().strip()
	out: list[DeviceRecord] = []
	for d in model.devices:
		if d.review_status == ReviewStatus.EXCLUDED:
			continue
		cls_sheet = SHEET_FOR_CLASS.get(d.device_class, "").lower().strip()
		if cls_sheet == want:
			out.append(d)
	return out


def _clear_device_rows(ws, sheet: TemplateSheet) -> None:
	"""Delete every cell value from `data_start_row` down, columns D onward.

	Metadata rows in cols B–C (UDT Type, Folder) must be preserved. Column D
	is the optional flag column the UFP template uses for `(not used)` markers
	on reserved-but-empty slots — these must be cleared so old flags don't
	bleed onto re-used rows. Headers (row 1) are never touched.
	"""
	# Always start at column D so the template's flag column is cleared too.
	first_data_col = 4
	last_data_col = max(
		ws.max_column,
		max((c.col_index for c in sheet.columns), default=4),
	)
	for r in range(sheet.data_start_row, ws.max_row + 1):
		for c in range(first_data_col, last_data_col + 1):
			ws.cell(row=r, column=c).value = None


def _write_device(ws, sheet: TemplateSheet, row: int, device: DeviceRecord) -> None:
	# Identity columns — fill what the template asks for. Register columns are
	# left blank (Plan §9A.5).
	for col in sheet.columns:
		if col.kind == ColumnKind.IDENTITY_SYSTEM:
			ws.cell(row=row, column=col.col_index).value = device.system.value
		elif col.kind == ColumnKind.IDENTITY_SYSTEM_NUMBER:
			ws.cell(row=row, column=col.col_index).value = device.system_number
		elif col.kind == ColumnKind.IDENTITY_NAME:
			ws.cell(row=row, column=col.col_index).value = ignition_name(device)
		elif col.kind == ColumnKind.IDENTITY_DESCRIPTION:
			ws.cell(row=row, column=col.col_index).value = device.description


def _write_sheet(wb: Workbook, sheet: TemplateSheet, devices: list[DeviceRecord]) -> int:
	ws = wb[sheet.name]
	_clear_device_rows(ws, sheet)
	# Sort devices so the output is deterministic and easy to scan: by system,
	# system number, then base name.
	devices = sorted(
		devices,
		key=lambda d: (d.system.value, d.system_number or 0, d.base_name.lower()),
	)
	row = sheet.data_start_row
	for d in devices:
		_write_device(ws, sheet, row, d)
		row += 1
	return len(devices)


# =============================================================================
# Tank sheet — driven by plant.tanks, not the DeviceModel
# =============================================================================
# All template Tank rows live under Mixing/System 1 in the UFP layout — the
# template puts every tank instance into the "Mixing" system group. Plant
# Configuration is the source for which tanks exist; the standard register
# pattern fills the address columns.
TANK_SHEET_SYSTEM = "Mixing"
TANK_SHEET_SYSTEM_NUMBER = 1


def _idle_flag_column(sheet: TemplateSheet) -> Optional[int]:
	"""Optional flag column the template uses for "(not used)" markers.

	Sits in column D in the UFP templates we've seen. We detect it positionally
	rather than by header text (the header for col D is empty).
	"""
	for col in sheet.columns:
		if col.col_index == 4:
			return col.col_index
	return 4 if sheet.columns else None


def _write_tank_row(
	ws, sheet: TemplateSheet, row: int, tank: TankRecord, slot: int
) -> None:
	"""Write one tank's row into the template: identity + standard registers.

	`slot` is the 1-based instance index used to compute MW addresses. Idle
	tanks still get a row and registers (the PLC keeps the slot allocated);
	we annotate the optional flag column to make them easy to spot.
	"""
	for col in sheet.columns:
		if col.kind == ColumnKind.IDENTITY_SYSTEM:
			ws.cell(row=row, column=col.col_index).value = TANK_SHEET_SYSTEM
		elif col.kind == ColumnKind.IDENTITY_SYSTEM_NUMBER:
			ws.cell(row=row, column=col.col_index).value = TANK_SHEET_SYSTEM_NUMBER
		elif col.kind == ColumnKind.IDENTITY_NAME:
			ws.cell(row=row, column=col.col_index).value = tank.tank_id
		elif col.kind == ColumnKind.IDENTITY_DESCRIPTION:
			ws.cell(row=row, column=col.col_index).value = (
				tank.chemical or ""
			)
		elif col.header in TANK_REGISTER_PATTERN:
			ws.cell(row=row, column=col.col_index).value = tank_register_for(col.header, slot)
		# Agitator.* and MixMan columns are intentionally left blank — they
		# are device-output mappings, not part of the standard tank block.

	# Mark idle tanks in the optional flag column (column D in UFP templates)
	# so the engineer can see them at a glance, mirroring how the source
	# template flags "(not used)" instances.
	if tank.is_idle:
		ws.cell(row=row, column=4).value = "(idle)"


def _write_tank_sheet(wb: Workbook, sheet: TemplateSheet, plant: PlantConfiguration) -> int:
	ws = wb[sheet.name]
	_clear_device_rows(ws, sheet)
	row = sheet.data_start_row
	for slot, tank in enumerate(plant.tanks, start=1):
		_write_tank_row(ws, sheet, row, tank, slot)
		row += 1
	return len(plant.tanks)


def export_io_list(
	template_path: str | Path,
	plant: PlantConfiguration,
	model: DeviceModel,
	output_dir: str | Path,
	tmap: Optional[TemplateMap] = None,
) -> Path:
	"""Render `model` into a copy of the UFP template and save it.

	Returns the path to the written file.
	"""
	if tmap is None:
		tmap = read_template(template_path)

	wb = load_template_workbook(template_path)

	tank_sheet = tmap.sheet_for_class(DeviceClass.TANK)
	for sheet in tmap.sheets.values():
		if sheet.name not in wb.sheetnames:
			continue
		# Tank sheet is special — its rows come from plant.tanks plus the UFP
		# standard register pattern (Plan §9A), not from the device model.
		if tank_sheet is not None and sheet.name == tank_sheet.name:
			_write_tank_sheet(wb, sheet, plant)
			continue
		devices = _devices_for_sheet(model, sheet)
		_write_sheet(wb, sheet, devices)

	out_dir = Path(output_dir)
	out_dir.mkdir(parents=True, exist_ok=True)
	out_path = out_dir / f"{_safe_site_name(plant)}_Ignition_IOList.xlsx"
	wb.save(out_path)
	return out_path
