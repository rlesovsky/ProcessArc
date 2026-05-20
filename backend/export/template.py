"""Template Reader (Plan §7) — build a TemplateMap from a UFP IO-list workbook.

For each sheet, classifies every column into one of three kinds:
  - identity        : filled from the Device Model (System, System Number, Name)
  - standard_register : pre-fillable from a UFP standard pattern (Plan §9A)
  - variable_register : left blank for the PLC programmers (the default)

Phase 1 safe default per Plan §9A.5: every register column is treated as
variable_register and left blank. The standard-register classification exists
so a future revision can flip individual columns to standard_register without
touching the exporter code.

The reader is forgiving about layout drift: an extra column it doesn't
recognize is kept (as variable_register / "leave alone"); a renamed sheet is
followed by its actual name; a missing sheet shows up in `warnings`.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook import Workbook

from backend.model.device import DeviceClass


# Maps DeviceClass → expected template sheet name. Matching is case-insensitive
# and lenient on whitespace; the engineer may have re-cased a sheet name.
SHEET_FOR_CLASS: dict[DeviceClass, str] = {
	DeviceClass.PUMP: "Pump",
	DeviceClass.VALVE: "Valve",
	DeviceClass.VFD_PUMP: "VFD Pump",
	DeviceClass.CONTROL_VALVE: "Control Valve",
	DeviceClass.TANK: "Tank",
}


class ColumnKind(str, Enum):
	IDENTITY_SYSTEM = "identity:system"
	IDENTITY_SYSTEM_NUMBER = "identity:system_number"
	IDENTITY_NAME = "identity:base_name"
	IDENTITY_DESCRIPTION = "identity:description"
	STANDARD_REGISTER = "standard_register"
	VARIABLE_REGISTER = "variable_register"


# Header text → column kind. Matched case-insensitively after stripping.
_IDENTITY_HEADER_ALIASES: dict[str, ColumnKind] = {
	"system name": ColumnKind.IDENTITY_SYSTEM,
	"system": ColumnKind.IDENTITY_SYSTEM,
	"system number": ColumnKind.IDENTITY_SYSTEM_NUMBER,
	"#": ColumnKind.IDENTITY_SYSTEM_NUMBER,
	"number": ColumnKind.IDENTITY_SYSTEM_NUMBER,
	"name": ColumnKind.IDENTITY_NAME,
	"tag name": ColumnKind.IDENTITY_NAME,
	"description": ColumnKind.IDENTITY_DESCRIPTION,
	"note": ColumnKind.IDENTITY_DESCRIPTION,
	"notes": ColumnKind.IDENTITY_DESCRIPTION,
}


@dataclass
class TemplateColumn:
	col_index: int  # 1-based, matches openpyxl
	header: str
	kind: ColumnKind


@dataclass
class TemplateSheet:
	name: str
	header_row: int
	udt_type: str = ""
	folder: str = ""
	# Rows reserved for metadata (UDT Type, Folder) that the exporter must
	# preserve. The data row range is everything below max(header_row,
	# metadata_rows) up to the existing max_row.
	metadata_rows: list[int] = field(default_factory=list)
	columns: list[TemplateColumn] = field(default_factory=list)

	def identity_column(self, kind: ColumnKind) -> Optional[TemplateColumn]:
		for c in self.columns:
			if c.kind == kind:
				return c
		return None

	@property
	def data_start_row(self) -> int:
		# Device rows start the row after the metadata block ends. In the UFP
		# templates seen so far this is row 2 (header on row 1, metadata in
		# cols B-C of rows 2-3 with device data on cols E+ starting row 2).
		return self.header_row + 1


@dataclass
class TemplateMap:
	source_path: Path
	sheets: dict[str, TemplateSheet] = field(default_factory=dict)
	warnings: list[str] = field(default_factory=list)

	def sheet_for_class(self, device_class: DeviceClass) -> Optional[TemplateSheet]:
		want = SHEET_FOR_CLASS.get(device_class, "").lower().strip()
		for name, sheet in self.sheets.items():
			if name.lower().strip() == want:
				return sheet
		return None


# =============================================================================
# Reader
# =============================================================================
def _classify_header(header: str) -> ColumnKind:
	"""Best-effort header → ColumnKind. Anything unrecognized is treated as
	variable_register (leave blank, don't touch). The exporter doesn't write
	unknown columns; the engineer's existing template content is preserved.
	"""
	key = (header or "").strip().lower()
	if not key:
		return ColumnKind.VARIABLE_REGISTER
	if key in _IDENTITY_HEADER_ALIASES:
		return _IDENTITY_HEADER_ALIASES[key]
	# Default: a register-shaped column we don't pre-fill in Phase 1 (§9A.5).
	return ColumnKind.VARIABLE_REGISTER


def _read_metadata(ws) -> tuple[str, str, list[int]]:
	"""Scan the first ~6 rows of cols B–C for 'UDT Type:' / 'Folder:'.

	Returns (udt_type, folder, rows_used). Rows used are the row numbers where
	metadata was found — the exporter must not clobber them.
	"""
	udt_type = ""
	folder = ""
	rows_used: list[int] = []
	for r in range(1, min(ws.max_row + 1, 8)):
		label = ws.cell(row=r, column=2).value
		value = ws.cell(row=r, column=3).value
		if label is None:
			continue
		key = str(label).strip().lower()
		if key.startswith("udt type"):
			udt_type = str(value or "").strip()
			rows_used.append(r)
		elif key.startswith("folder"):
			folder = str(value or "").strip()
			rows_used.append(r)
	return udt_type, folder, rows_used


def _read_sheet(ws) -> TemplateSheet:
	# Header is row 1 in every UFP template seen so far. Identity columns sit
	# from column E (index 5) onward. We classify every non-empty header cell.
	header_row = 1
	columns: list[TemplateColumn] = []
	for c in range(1, ws.max_column + 1):
		header_val = ws.cell(row=header_row, column=c).value
		if header_val is None:
			continue
		text = str(header_val)
		columns.append(TemplateColumn(col_index=c, header=text, kind=_classify_header(text)))

	udt_type, folder, metadata_rows = _read_metadata(ws)
	return TemplateSheet(
		name=ws.title,
		header_row=header_row,
		udt_type=udt_type,
		folder=folder,
		metadata_rows=metadata_rows,
		columns=columns,
	)


def read_template(path: str | Path) -> TemplateMap:
	"""Open the UFP IO-list template and build a TemplateMap.

	Reads in data-only mode (formula values resolved) since we're inspecting
	structure, not formulas. The exporter loads the file again with formulas
	preserved when it writes — see `io_list.export_io_list`.
	"""
	p = Path(path)
	wb = openpyxl.load_workbook(p, data_only=True)

	tmap = TemplateMap(source_path=p)
	for sn in wb.sheetnames:
		ws = wb[sn]
		# Skip the Config sheet — it carries site-level identity, not device rows.
		if sn.strip().lower() == "config":
			continue
		tmap.sheets[sn] = _read_sheet(ws)

	# Warn about expected sheets that are missing entirely.
	expected = {name.lower(): name for name in SHEET_FOR_CLASS.values()}
	present = {n.lower() for n in tmap.sheets}
	for k, original in expected.items():
		if k not in present:
			tmap.warnings.append(f"Expected sheet '{original}' not found in template.")

	# Warn about sheets we found but couldn't map (e.g. a renamed sheet).
	for n, sheet in tmap.sheets.items():
		if not any(n.lower().strip() == w.lower().strip() for w in SHEET_FOR_CLASS.values()):
			tmap.warnings.append(
				f"Sheet '{n}' did not match any known UDT type — devices won't be written there."
			)

	return tmap


def load_template_workbook(path: str | Path) -> Workbook:
	"""Open the template for writing — formulas preserved, no data-only."""
	return openpyxl.load_workbook(Path(path))
