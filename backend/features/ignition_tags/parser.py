"""Workbook parser for the Ignition Tag Builder.

Reads the user-supplied .xlsx with openpyxl and produces a typed
representation that the builder consumes. This module is a faithful
port of the row/column iteration logic in the reference Jython script
(see backend/features/ignition_tags/README.md for the source).

==============================================================================
Indexing — read this before editing
==============================================================================
The reference Jython uses Apache POI, which is 0-indexed for both rows
and columns. openpyxl is 1-indexed. The two conventions overlap exactly
once (the user-visible "spreadsheet row/column") and that is the
convention we use in all user-facing error messages.

    cell C2  →  POI row=1, col=2  →  openpyxl row=2, column=3
    cell C3  →  POI row=2, col=2  →  openpyxl row=3, column=3
    cell E1  →  POI row=0, col=4  →  openpyxl row=1, column=5

`get_cell_value(file, row=1, col=2)` in the Jython reads cell **C2**
(POI col=2 is the third column, which is C). The provider name, site
name, UDT type id, and folder all live in column C (column B holds
human-readable labels like "Tag Provider:" / "UDT Type:").

The data table on each sheet starts at zero-indexed row 0 (spreadsheet
row 1, openpyxl row 1) and zero-indexed column 4 (spreadsheet column E,
openpyxl column 5).
==============================================================================
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schema import ValidationIssue

# The three columns that must appear in the first three positions of
# every data sheet, in any order. The Jython skips the first three
# positions and looks the values up by name, so order is not enforced.
REQUIRED_COLUMNS: tuple[str, ...] = ("Name", "System Name", "System Number")

# Spreadsheet column letter where the data table header row begins
# (zero-indexed column 4 in the Jython, which is openpyxl column 5).
DATA_FIRST_COLUMN = 5

# Spreadsheet row where the data table header row sits
# (zero-indexed row 0 in the Jython, which is openpyxl row 1).
DATA_HEADER_ROW = 1

# Column where the per-sheet config values live (provider/site in the
# header sheet; UDT type id and folder in each data sheet). The Jython
# reads POI col=2, which is spreadsheet column C, openpyxl column 3.
CONFIG_VALUE_COLUMN = 3


@dataclass
class SheetData:
	"""A parsed data sheet — one UDT type per sheet."""

	sheet_name: str
	udt_type: str
	folder: str
	headers: list[str]
	rows: list[dict[str, Any]]  # one dict per data row keyed by header name


@dataclass
class ParsedWorkbook:
	"""The result of parsing the entire workbook.

	`sheets` is empty if the workbook has no usable data sheets (this
	is reported as an error and the request will be rejected; the
	parser collects rather than raises).
	"""

	provider: str | None
	site: str | None
	sheets: list[SheetData] = field(default_factory=list)
	issues: list[ValidationIssue] = field(default_factory=list)


def _coerce_numeric(value: Any) -> Any:
	"""Match the Jython numeric-coercion behavior for cell values.

	A float that round-trips exactly as int becomes int. Everything
	else passes through unchanged.
	"""
	if isinstance(value, float) and value == int(value):
		return int(value)
	return value


def _read_cell(ws: Worksheet, row: int, column: int) -> Any:
	"""Read a cell using openpyxl row/column (both 1-indexed).

	Returns the coerced value, or None if the cell is empty or holds an
	empty string. Strings are stripped — leading/trailing whitespace in
	a header would silently break the parser's lookup otherwise.
	"""
	cell = ws.cell(row=row, column=column)
	value = cell.value
	if value is None:
		return None
	if isinstance(value, str):
		stripped = value.strip()
		return stripped if stripped else None
	return _coerce_numeric(value)


def _parse_header_sheet(ws: Worksheet, issues: list[ValidationIssue]) -> tuple[str | None, str | None]:
	"""Sheet 0 — read provider (C2) and site (C3)."""
	provider = _read_cell(ws, 2, CONFIG_VALUE_COLUMN)
	site = _read_cell(ws, 3, CONFIG_VALUE_COLUMN)
	if provider is None:
		issues.append(
			ValidationIssue(
				severity="error",
				code="header.missing_provider",
				message="Sheet 0 cell C2 (tag provider) is blank.",
				sheet=ws.title,
				row=2,
				column="C",
			)
		)
	if site is None:
		issues.append(
			ValidationIssue(
				severity="error",
				code="header.missing_site",
				message="Sheet 0 cell C3 (site name) is blank.",
				sheet=ws.title,
				row=3,
				column="C",
			)
		)
	return (
		None if provider is None else str(provider),
		None if site is None else str(site),
	)


def _read_data_headers(ws: Worksheet) -> list[str]:
	"""Read the header row of a data sheet.

	Walks from `DATA_FIRST_COLUMN` rightward until the first blank
	header cell. Mirrors the Jython `table_to_dataset` auto-detect of
	`lastCol`.
	"""
	headers: list[str] = []
	column = DATA_FIRST_COLUMN
	while True:
		value = _read_cell(ws, DATA_HEADER_ROW, column)
		if value is None:
			break
		headers.append(str(value))
		column += 1
	return headers


def _name_column_index(headers: list[str]) -> int | None:
	"""Position of the `Name` column in `headers`, or None if absent.

	The Jython terminates the data table at the first row whose `Name`
	cell is blank — but `Name` can appear in any of the first three
	positions, so we have to look it up.
	"""
	try:
		return headers.index("Name")
	except ValueError:
		return None


def _read_data_rows(
	ws: Worksheet, headers: list[str], name_column_index: int
) -> list[dict[str, Any]]:
	"""Read data rows for a sheet, terminating at the first blank `Name`.

	Mirrors the Jython auto-detect of `lastRow` (first blank in the
	first data column).
	"""
	rows: list[dict[str, Any]] = []
	# Data rows start at openpyxl row 2 (zero-indexed row 1 in the Jython).
	row = DATA_HEADER_ROW + 1
	while True:
		name_value = _read_cell(ws, row, DATA_FIRST_COLUMN + name_column_index)
		if name_value is None:
			break
		record: dict[str, Any] = {}
		for idx, header in enumerate(headers):
			record[header] = _read_cell(ws, row, DATA_FIRST_COLUMN + idx)
		record["__row__"] = row
		rows.append(record)
		row += 1
	return rows


def _parse_data_sheet(ws: Worksheet, issues: list[ValidationIssue]) -> SheetData | None:
	"""Parse one UDT data sheet.

	Returns None if the sheet has a fatal structural problem; in that
	case the corresponding error is added to `issues`.
	"""
	udt_type = _read_cell(ws, 2, CONFIG_VALUE_COLUMN)
	folder = _read_cell(ws, 3, CONFIG_VALUE_COLUMN)
	if udt_type is None:
		issues.append(
			ValidationIssue(
				severity="error",
				code="sheet.missing_udt_type",
				message=f"Sheet {ws.title!r} cell C2 (UDT type id) is blank.",
				sheet=ws.title,
				row=2,
				column="C",
			)
		)
	if folder is None:
		issues.append(
			ValidationIssue(
				severity="error",
				code="sheet.missing_folder",
				message=f"Sheet {ws.title!r} cell C3 (destination folder) is blank.",
				sheet=ws.title,
				row=3,
				column="C",
			)
		)
	if udt_type is None or folder is None:
		return None

	headers = _read_data_headers(ws)

	# Required-column check: Name, System Name, and System Number must
	# all appear in the first 3 positions. The Jython doesn't enforce
	# order — it skips the first 3 positionally and looks the three
	# values up by name in the dataset.
	first_three = set(headers[: len(REQUIRED_COLUMNS)])
	missing = [c for c in REQUIRED_COLUMNS if c not in first_three]
	if missing:
		actual = headers[: len(REQUIRED_COLUMNS)] or ["(no headers)"]
		issues.append(
			ValidationIssue(
				severity="error",
				code="sheet.missing_required_column",
				message=(
					f"Sheet {ws.title!r}: the first 3 header columns must include "
					f"{', '.join(repr(c) for c in REQUIRED_COLUMNS)} (any order); "
					f"missing {', '.join(repr(c) for c in missing)}. "
					f"Found: {', '.join(repr(h) for h in actual)}."
				),
				sheet=ws.title,
				row=DATA_HEADER_ROW,
			)
		)
		return None

	name_col_idx = _name_column_index(headers)
	# `name_col_idx` cannot be None here — the missing-column check above
	# already verified `Name` is present in the first three. Asserting
	# for the type checker.
	assert name_col_idx is not None

	rows = _read_data_rows(ws, headers, name_col_idx)
	if not rows:
		# Matches the Jython, which silently iterates zero rows for an
		# empty sheet. Surfaced as a warning so the engineer notices the
		# sheet was effectively skipped — without failing the workbook.
		issues.append(
			ValidationIssue(
				severity="warning",
				code="sheet.no_data_rows",
				message=(
					f"Sheet {ws.title!r} has zero data rows — no instances "
					"will be generated for this sheet."
				),
				sheet=ws.title,
				row=DATA_HEADER_ROW + 1,
			)
		)
		# Return the sheet anyway so it appears in the manifest with an
		# empty row list; the builder will simply skip it.
		return SheetData(
			sheet_name=ws.title,
			udt_type=str(udt_type),
			folder=str(folder),
			headers=headers,
			rows=[],
		)

	return SheetData(
		sheet_name=ws.title,
		udt_type=str(udt_type),
		folder=str(folder),
		headers=headers,
		rows=rows,
	)


def parse_workbook(file_bytes: bytes) -> ParsedWorkbook:
	"""Parse the user-uploaded xlsx into a typed `ParsedWorkbook`.

	Read-only and data-only (formulas resolved to their cached values).
	The parser collects every issue rather than raising — the router
	decides whether to fail the request based on `issues`.
	"""
	wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
	sheet_names = wb.sheetnames
	issues: list[ValidationIssue] = []

	if len(sheet_names) < 2:
		issues.append(
			ValidationIssue(
				severity="error",
				code="workbook.too_few_sheets",
				message=(
					"Workbook must have at least 2 sheets: a header sheet "
					"(Sheet 0) and one or more data sheets."
				),
			)
		)
		return ParsedWorkbook(provider=None, site=None, sheets=[], issues=issues)

	header_ws = wb[sheet_names[0]]
	provider, site = _parse_header_sheet(header_ws, issues)

	parsed_sheets: list[SheetData] = []
	for name in sheet_names[1:]:
		ws = wb[name]
		parsed = _parse_data_sheet(ws, issues)
		if parsed is not None:
			parsed_sheets.append(parsed)

	wb.close()

	return ParsedWorkbook(
		provider=provider,
		site=site,
		sheets=parsed_sheets,
		issues=issues,
	)
