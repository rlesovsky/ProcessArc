"""Parse a customer write-up workbook into a structured ``ParsedSource``.

The write-up format we target is the "Graphics and Tables" style xlsx
the plant team supplies for each commissioning project. Known sheets:

  - ``Chemical``                — flow-meter table (chem, meter, K-factor, make/model)
  - ``Cylinder 1 Sequencing``   — free-text step-by-step narrative
  - ``Cylinder 2 Sequencing``   — same shape as Cyl 1
  - ``Mix Sequencing``          — free-text mix sequence narrative
  - ``Cylinder 1 Treat Graphic`` — notes (deviations from canonical graphic)
  - ``Cylinder 2 Treat Graphic`` — notes
  - ``Mix Graphic``             — notes
  - ``Plant Info``              — short list of plant-level facts
  - ``Operators`` / ``Tank Info`` / ``Tank Default Info``
      — usually empty placeholders today, but we read them if populated

We're permissive: unknown sheets are recorded in ``unknown_sheets`` and
ignored. Missing-but-expected sheets become parser-level warnings, not
errors — the builder can still produce a useful workbook from a
partial source.
"""

from __future__ import annotations

import io
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schema import (
    FlowMeter,
    GraphicNote,
    ParsedSource,
    PlantInfo,
    SequenceNote,
    _coerce_str,
)


# Canonical step names we recognize in the sequence narratives. Order
# matters: when scanning row-by-row, hitting one of these starts a new
# step bucket.
KNOWN_STEPS = (
    "Initial Vacuum",
    "Fill",
    "Raise Pressure",
    "Pressure Relief",
    "Pressure",
    "Empty",
    "Final Vacuum",
    "Final Empty",
    "Finish",
)


def parse_source(xlsx_bytes: bytes) -> ParsedSource:
    """Top-level: read the bytes, dispatch each sheet to its handler."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    parsed = ParsedSource()
    recognized: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        s = sheet_name.strip()
        if s.lower() == "chemical":
            parsed.flow_meters.extend(_parse_chemical(ws))
            recognized.add(sheet_name)
        elif s.lower().startswith("cylinder 1 sequencing"):
            parsed.sequence_notes.extend(_parse_sequencing(ws, cylinder=1))
            recognized.add(sheet_name)
        elif s.lower().startswith("cylinder 2 sequencing"):
            parsed.sequence_notes.extend(_parse_sequencing(ws, cylinder=2))
            recognized.add(sheet_name)
        elif s.lower().startswith("mix sequencing"):
            parsed.sequence_notes.extend(_parse_sequencing(ws, cylinder=None))
            recognized.add(sheet_name)
        elif s.lower().startswith("cylinder 1 treat graphic"):
            parsed.graphic_notes.append(_parse_graphic(ws, section="Cylinder 1"))
            recognized.add(sheet_name)
        elif s.lower().startswith("cylinder 2 treat graphic"):
            parsed.graphic_notes.append(_parse_graphic(ws, section="Cylinder 2"))
            recognized.add(sheet_name)
        elif s.lower().startswith("mix graphic"):
            parsed.graphic_notes.append(_parse_graphic(ws, section="Mix"))
            recognized.add(sheet_name)
        elif s.lower() == "plant info":
            parsed.plant_info.plant_facts.extend(_parse_plant_info(ws))
            recognized.add(sheet_name)
        elif s.lower() == "operators":
            parsed.plant_info.operator_names.extend(_parse_single_column(ws))
            recognized.add(sheet_name)
        elif s.lower() in ("tank info", "tank default info", "cylinder voids"):
            parsed.plant_info.tank_notes.extend(_parse_single_column(ws))
            recognized.add(sheet_name)

    parsed.unknown_sheets = [s for s in wb.sheetnames if s not in recognized]
    return parsed


# ─── Per-sheet handlers ─────────────────────────────────────────────────────

def _parse_chemical(ws: Worksheet) -> list[FlowMeter]:
    """Chemical sheet → flow-meter rows.

    The 'Flow Meter Info' header sits at row 17 with the data rows
    starting at row 19. Columns: A=Chem, B=Meter Description, C=K-Factor,
    D=Make/Model.
    """
    meters: list[FlowMeter] = []
    # We scan a fixed range; sheets that don't match the layout simply
    # yield no rows.
    for row in ws.iter_rows(min_row=19, max_row=40, values_only=True):
        if not row or row[0] is None:
            continue
        chem = _coerce_str(row[0])
        if not chem:
            continue
        # Cut-off keywords — the sheet often has follow-on sections
        # ('QC Factors', 'Set MCA strip factor...') after the meter
        # table. Bail when we see those.
        if chem.lower().startswith(("qc factors", "set ", "the ")):
            break
        meter_desc = _coerce_str(row[1] if len(row) > 1 else None)
        k = row[2] if len(row) > 2 else None
        make = _coerce_str(row[3] if len(row) > 3 else None)
        meters.append(FlowMeter(
            chemical=chem,
            meter_description=meter_desc,
            k_factor=k if isinstance(k, (int, float)) else _coerce_k(k),
            make_model=make,
        ))
    return meters


def _coerce_k(v) -> float | int | None:
    """Best-effort K-factor parse — sources sometimes use strings."""
    if v is None:
        return None
    s = _coerce_str(v)
    if not s:
        return None
    try:
        return int(s) if s.isdigit() else float(s)
    except (TypeError, ValueError):
        return None


def _parse_sequencing(ws: Worksheet, cylinder: int | None) -> list[SequenceNote]:
    """Cyl 1/2 / Mix Sequencing → SequenceNote per recognized step.

    The narrative is laid out as plain rows in col A (sometimes A-Y but
    only A reliably has content). A row that exactly matches one of
    KNOWN_STEPS (case-insensitive) opens a new step bucket; subsequent
    non-blank rows accumulate as that step's notes until the next match.
    """
    notes: list[SequenceNote] = []
    current: SequenceNote | None = None
    for row in ws.iter_rows(values_only=True):
        first = _coerce_str(row[0] if row else None)
        if not first:
            continue
        # Sequence-step header?
        match = _match_step(first)
        if match is not None:
            if current is not None:
                notes.append(current)
            current = SequenceNote(cylinder=cylinder, step_name=match)
            continue
        # Otherwise accumulate as a note on the current step (if any).
        if current is not None:
            current.notes.append(first)
    if current is not None:
        notes.append(current)
    return notes


def _match_step(text: str) -> str | None:
    """Return the canonical step name if ``text`` looks like a step header."""
    t = text.strip()
    # Reject obvious non-headers (long rows, bullet text, etc.) by length.
    if len(t) > 60:
        return None
    lowered = t.lower()
    for step in KNOWN_STEPS:
        if lowered == step.lower():
            return step
    return None


def _parse_graphic(ws: Worksheet, section: str) -> GraphicNote:
    """*Graphic* sheet → free-text notes for the given section.

    Layout in Union City: header 'Notes' in col P/R/O (varies), then
    one note per row. We just collect every non-empty cell from rows
    1-20 across all columns and de-dupe.
    """
    seen: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for cell in row or ():
            s = _coerce_str(cell)
            if not s:
                continue
            if s.lower() == "notes":
                continue
            if s in seen:
                continue
            seen.append(s)
    return GraphicNote(section=section, notes=seen)


def _parse_plant_info(ws: Worksheet) -> list[str]:
    """Plant Info → list of short factual lines."""
    return _parse_single_column(ws)


def _parse_single_column(ws: Worksheet) -> list[str]:
    """Generic helper for single-column sheets — Operators, Tank Info, etc."""
    out: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for cell in row or ():
            s = _coerce_str(cell)
            if not s:
                continue
            out.append(s)
            break  # One value per row at most.
    return out
