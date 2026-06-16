"""Apply a ``ParsedSource`` to the CommWKBK template.

Four mapping rules, in the order they run:

1. ``Chemical`` → ``Tank And Chem Number``
   Adds two annotation columns (M = Source K-Factor, N = Source Meter) and
   fills them for rows in the MIX System 1 block whose chemical name
   matches a source row. Existing template K-Factors are NEVER overwritten.

2. ``Cyl/Mix Sequencing`` → Sign-Off COMMENTS columns
   For each sequence step in the source, locate the matching ``STEP n -
   <name>`` header in ``Treat Sequence Sign Off`` (Cyl 1 only — the
   template doesn't have a Cyl 2 section) or ``Mix Sequence Sign Off``,
   and drop the source's note text into the COMMENTS column (col E) of
   that header row.

3. ``*Treat Graphic / Mix Graphic`` notes → COMMENTS at the section
   header of the matching sign-off sheet (Cyl 1 / Mix sections).

4. ``Plant Info / Operators / Tank Info`` → Network Schema notes column
   (col G). Each fact becomes its own row at the bottom of the sheet.

The builder is strictly additive — it never overwrites a cell that
already holds content. Conflicts are recorded in the ``BuildReport`` so
the UI can show them.
"""

from __future__ import annotations

import io
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schema import BuildReport, ChangeLogEntry, ParsedSource, _coerce_str


# Annotation columns appended to Tank And Chem Number. Cols A–L are
# already in use (B Tank No., C Description, F Chem No., G Description,
# H Card, I Channel, J K Factor, K Unit from SCADA). M and N are free.
TANK_COL_SRC_K = "M"
TANK_COL_SRC_METER = "N"


def build_workbook(
    parsed: ParsedSource,
    template_bytes: bytes,
    template_name: str = "default_commissioning_workbook.xlsx",
) -> tuple[bytes, BuildReport]:
    """Apply ``parsed`` to ``template_bytes``; return (new xlsx bytes, report)."""
    wb = load_workbook(io.BytesIO(template_bytes))
    report = BuildReport(template_name=template_name)
    report.source_sheets_seen = (
        ["Chemical"] * bool(parsed.flow_meters)
        + (["Cylinder Sequencing"] if any(
            n.cylinder is not None for n in parsed.sequence_notes) else [])
        + (["Mix Sequencing"] if any(
            n.cylinder is None for n in parsed.sequence_notes) else [])
        + [g.section + " Graphic" for g in parsed.graphic_notes]
        + (["Plant Info"] if parsed.plant_info.plant_facts else [])
    )
    # Strip empty entries and dedupe while preserving order.
    seen: list[str] = []
    for s in report.source_sheets_seen:
        if s and s not in seen:
            seen.append(s)
    report.source_sheets_seen = seen

    if parsed.unknown_sheets:
        report.warnings.append(
            f"Ignored {len(parsed.unknown_sheets)} unrecognized source sheets: "
            + ", ".join(parsed.unknown_sheets)
        )

    _apply_chemical(wb, parsed, report)
    _apply_sequence(wb, parsed, report)
    _apply_graphic_notes(wb, parsed, report)
    _apply_plant_info(wb, parsed, report)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report


# ─── Generic write helper that records every touch ──────────────────────────

def _safe_set(
    ws: Worksheet,
    cell: str,
    value,
    reason: str,
    report: BuildReport,
    *,
    allow_overwrite: bool = False,
) -> None:
    c = ws[cell]
    before = "" if c.value is None else str(c.value)
    if before.strip() and not allow_overwrite:
        report.changes.append(ChangeLogEntry(
            sheet=ws.title, cell=cell, before=before,
            after=str(value), reason=reason, conflict=True,
        ))
        return
    c.value = value
    report.changes.append(ChangeLogEntry(
        sheet=ws.title, cell=cell, before=before,
        after=str(value), reason=reason,
    ))


# ─── Mapping 1: Chemical → Tank And Chem Number ─────────────────────────────

def _apply_chemical(wb: Workbook, parsed: ParsedSource, report: BuildReport) -> None:
    if not parsed.flow_meters:
        return
    if "Tank And Chem Number" not in wb.sheetnames:
        report.warnings.append(
            "Template has no 'Tank And Chem Number' sheet; skipping flow-meter cross-check."
        )
        return
    ws = wb["Tank And Chem Number"]

    # Index source meters by lowercased chemical name. Multiple meters
    # per chemical (e.g. MCA-1 / MCA-2) are concatenated in display.
    by_chem: dict[str, list] = {}
    for fm in parsed.flow_meters:
        by_chem.setdefault(fm.chemical.strip().lower(), []).append(fm)

    # Headers + banner above the new columns. Banner sits at row 2 to
    # mirror the "ALL TANKS / MIX System 1" header block already in r2.
    _safe_set(ws, f"{TANK_COL_SRC_K}2",
              "Cross-check from Graphics+Tables → Chemical sheet. "
              "Existing K Factors in col J were NOT overwritten.",
              "Annotation banner", report)
    _safe_set(ws, f"{TANK_COL_SRC_K}4",
              "Source K-Factor (Graphics+Tables)",
              "New annotation column", report)
    _safe_set(ws, f"{TANK_COL_SRC_METER}4",
              "Source Meter (Make / Model)",
              "New annotation column", report)

    # MIX System 1 data rows: 5..15 in the canonical template. Walk and
    # match by col G (chemical name).
    matched = 0
    for r in range(5, 16):
        chem_cell = ws.cell(row=r, column=7)  # col G
        if chem_cell.value is None:
            continue
        chem_name = _coerce_str(chem_cell.value).lower()
        candidates = by_chem.get(chem_name)
        if not candidates:
            continue
        k_str = " / ".join(
            str(c.k_factor) for c in candidates if c.k_factor is not None
        )
        desc_str = " / ".join(
            _format_meter(c)
            for c in candidates
            if (c.meter_description or c.make_model)
        )
        if k_str:
            _safe_set(ws, f"{TANK_COL_SRC_K}{r}", k_str,
                      f"Source K-Factor for chemical {chem_cell.value!r}",
                      report)
        if desc_str:
            _safe_set(ws, f"{TANK_COL_SRC_METER}{r}", desc_str,
                      f"Source meter make/model for chemical {chem_cell.value!r}",
                      report)
        matched += 1
    report.flow_meters_matched = matched


def _format_meter(fm) -> str:
    """Render a FlowMeter as 'description (make/model)', degrading gracefully
    when one side is missing. Avoids the dangling-paren bug that an
    over-eager strip("()") was causing on values like 'azbil (magtech)'."""
    desc = fm.meter_description.strip()
    make = fm.make_model.strip()
    if desc and make:
        return f"{desc} ({make})"
    return desc or make


# ─── Mapping 2: Sequence notes → COMMENTS columns ───────────────────────────

# Map source step name → (template sheet, search-for-text). The template
# uses "STEP n - <name>" headers; we search col A for an exact match.
TREAT_SHEET = "Treat Sequence Sign Off"
MIX_SHEET = "Mix Sequence Sign Off"


def _apply_sequence(wb: Workbook, parsed: ParsedSource, report: BuildReport) -> None:
    """Apply sequence notes to COMMENTS columns.

    The template has ONE Treat Sequence Sign Off — but the source has
    separate Cyl 1 / Cyl 2 sequences. Group notes by (sheet, row)
    before writing so both cylinders land in the same cell with clear
    section dividers, instead of Cyl 1 winning and Cyl 2 conflicting.
    """
    # Group notes by (sheet_name, step_name). Preserve insertion order
    # so Cyl 1 appears before Cyl 2.
    grouped: dict[tuple[str, str], list] = {}
    for note in parsed.sequence_notes:
        if not note.notes:
            continue
        sheet_name = MIX_SHEET if note.cylinder is None else TREAT_SHEET
        key = (sheet_name, note.step_name)
        grouped.setdefault(key, []).append(note)

    attached = 0
    for (sheet_name, step_name), notes in grouped.items():
        if sheet_name not in wb.sheetnames:
            report.warnings.append(
                f"Template has no '{sheet_name}' sheet; skipping {step_name!r}."
            )
            continue
        ws = wb[sheet_name]
        row = _find_step_row(ws, step_name)
        if row is None:
            report.warnings.append(
                f"No 'STEP n - {step_name}' row in {sheet_name!r}; "
                f"skipped {sum(len(n.notes) for n in notes)} note(s) "
                f"for {step_name}."
            )
            continue

        sections: list[str] = []
        for note in notes:
            label = (f"Cyl {note.cylinder}" if note.cylinder is not None else "Mix")
            body = " | ".join(note.notes[:6])
            if len(note.notes) > 6:
                body += f" ... (+{len(note.notes) - 6} more)"
            sections.append(f"[{label}] {body}")
        text = "From source — " + "  //  ".join(sections)

        _safe_set(ws, f"E{row}", text,
                  f"Sequence notes for {step_name} from source",
                  report)
        attached += len(notes)
    report.sequence_notes_attached = attached


def _find_step_row(ws: Worksheet, step_name: str) -> int | None:
    """Return the row number whose col A starts with 'STEP' and contains step_name."""
    target = step_name.lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        v = _coerce_str(cell.value)
        if not v:
            continue
        if v.lower().startswith("step ") and target in v.lower():
            return cell.row
    return None


# ─── Mapping 3: Graphic notes → COMMENTS at section headers ─────────────────

def _apply_graphic_notes(wb: Workbook, parsed: ParsedSource, report: BuildReport) -> None:
    attached = 0
    for graphic in parsed.graphic_notes:
        if not graphic.notes:
            continue
        # Find the right sheet + the section header row to anchor the note.
        # We treat:
        #  - Cylinder 1 → first "STEP 1 - Initial Vacuum" in Treat Sequence Sign Off
        #  - Cylinder 2 → currently unsupported (template doesn't have a Cyl2 section)
        #  - Mix → first row of Mix Sequence Sign Off
        if graphic.section == "Cylinder 1":
            sheet_name, anchor = TREAT_SHEET, "STEP 1"
        elif graphic.section == "Mix":
            sheet_name, anchor = MIX_SHEET, "MIX 1 SEQUENCE"
        else:
            report.warnings.append(
                f"Graphic notes for {graphic.section!r} ignored — no template anchor."
            )
            continue
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        anchor_row = _find_first_row_starting_with(ws, anchor)
        if anchor_row is None:
            report.warnings.append(
                f"Couldn't find anchor {anchor!r} in {sheet_name!r}; "
                f"graphic notes for {graphic.section} skipped."
            )
            continue
        text = f"Graphic notes ({graphic.section}): " + " | ".join(graphic.notes[:4])
        if len(graphic.notes) > 4:
            text += f" ... (+{len(graphic.notes) - 4} more)"
        _safe_set(ws, f"E{anchor_row}", text,
                  f"Graphic notes for {graphic.section} from source",
                  report)
        attached += 1
    report.graphic_notes_attached = attached


def _find_first_row_starting_with(ws: Worksheet, needle: str) -> int | None:
    n = needle.lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = _coerce_str(row[0].value)
        if v.lower().startswith(n):
            return row[0].row
    return None


# ─── Mapping 4: Plant facts → Network Schema notes ──────────────────────────

def _apply_plant_info(wb: Workbook, parsed: ParsedSource, report: BuildReport) -> None:
    facts: list[str] = []
    facts.extend(parsed.plant_info.plant_facts)
    if parsed.plant_info.operator_names:
        facts.append("Operators: " + ", ".join(parsed.plant_info.operator_names))
    if parsed.plant_info.tank_notes:
        facts.extend(f"Tank note: {n}" for n in parsed.plant_info.tank_notes)
    if not facts:
        return
    if "Network Schema" not in wb.sheetnames:
        report.warnings.append(
            "Template has no 'Network Schema' sheet; plant-level facts dropped."
        )
        return
    ws = wb["Network Schema"]
    start_row = ws.max_row + 1
    for i, fact in enumerate(facts):
        r = start_row + i
        _safe_set(ws, f"A{r}", "Plant fact",
                  "Plant fact from source", report)
        _safe_set(ws, f"G{r}", fact,
                  f"From source plant info: {fact[:60]}",
                  report)
    report.plant_facts_attached = len(facts)
