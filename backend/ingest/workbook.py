"""Workbook Ingester (Plan Stage 1).

Opens an UFP sequence workbook and classifies every sheet, without making any
extraction decisions. Downstream modules (Plant Configuration Discovery, Table
Extractor, Prose Extractor) consume the classified result.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional

import openpyxl


_CYLINDER_SEQ_RE = re.compile(r"^cylinder\s+(\d+)\s+sequencing$", re.IGNORECASE)
_MIX_SEQ_RE = re.compile(r"^(?:(.+?)\s+)?mix(?:ing)?\s+sequencing$", re.IGNORECASE)
_GRAPHIC_RE = re.compile(r"\bgraphic(s)?\b", re.IGNORECASE)
_CHEM_TANK_RE = re.compile(r"chemical\s+and\s+tank", re.IGNORECASE)


class SheetKind(str, Enum):
    CYLINDER_SEQUENCE = "cylinder_sequence"
    MIX_SEQUENCE = "mix_sequence"
    GRAPHIC = "graphic"
    CHEMICAL_AND_TANK = "chemical_and_tank"
    OTHER = "other"


@dataclass
class IngestedSheet:
    name: str
    kind: SheetKind
    rows: list[tuple]  # row values (cells -> python values, None for empty)
    cylinder_number: Optional[int] = None
    mix_label: Optional[str] = None  # e.g. "ECO", "MCA", "" for unified

    def text_lines(self, max_cols: int = 4) -> list[str]:
        """Flatten the prose part of the sheet to a list of non-empty lines.
        Used by the prose extractor — strips empties and stitches A..max_cols.
        """
        out: list[str] = []
        for row in self.rows:
            pieces = [str(c).strip() for c in row[:max_cols] if c is not None and str(c).strip()]
            if pieces:
                out.append(" — ".join(pieces) if len(pieces) > 1 else pieces[0])
        return out


@dataclass
class IngestedWorkbook:
    path: Path
    filename: str
    sheets: list[IngestedSheet] = field(default_factory=list)

    def by_kind(self, kind: SheetKind) -> list[IngestedSheet]:
        return [s for s in self.sheets if s.kind == kind]

    def find(self, kind: SheetKind) -> Optional[IngestedSheet]:
        for s in self.sheets:
            if s.kind == kind:
                return s
        return None


def classify_sheet(name: str) -> tuple[SheetKind, Optional[int], Optional[str]]:
    m = _CYLINDER_SEQ_RE.match(name.strip())
    if m:
        return SheetKind.CYLINDER_SEQUENCE, int(m.group(1)), None

    m = _MIX_SEQ_RE.match(name.strip())
    if m:
        label = (m.group(1) or "").strip()
        return SheetKind.MIX_SEQUENCE, None, label

    if _CHEM_TANK_RE.search(name):
        return SheetKind.CHEMICAL_AND_TANK, None, None

    if _GRAPHIC_RE.search(name):
        return SheetKind.GRAPHIC, None, None

    return SheetKind.OTHER, None, None


def ingest_workbook(path: str | Path) -> IngestedWorkbook:
    p = Path(path)
    wb = openpyxl.load_workbook(p, data_only=True, read_only=False)

    result = IngestedWorkbook(path=p, filename=p.name)

    for name in wb.sheetnames:
        kind, cyl, mix = classify_sheet(name)
        ws = wb[name]

        # Skip body load for graphic sheets — they hold images, no values
        if kind == SheetKind.GRAPHIC:
            rows: list[tuple] = []
        else:
            rows = [tuple(row) for row in ws.iter_rows(values_only=True)]

        result.sheets.append(
            IngestedSheet(
                name=name,
                kind=kind,
                rows=rows,
                cylinder_number=cyl,
                mix_label=mix,
            )
        )

    return result
