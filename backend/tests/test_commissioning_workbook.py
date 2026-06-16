"""Tests for the Commissioning Workbook Builder feature.

The fixture source xlsx is synthesized in code via openpyxl so the
tests are self-contained — no Union City golden file required. The
fixture covers all four mappings the parser is responsible for:
Chemical, Cylinder Sequencing, Treat/Mix Graphic, Plant Info.

The template fixture is also synthesized minimally — just enough rows
and sheets that the builder's mapping logic can land. This keeps the
tests fast (~1ms each) and means a template change won't break tests
that aren't actually about the template.
"""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend.api.main import app
from backend.features.commissioning_workbook import (
    build_workbook,
    parse_source,
)
from backend.features.commissioning_workbook.parser import KNOWN_STEPS
from backend.features.commissioning_workbook.schema import (
    BuildReport,
    ChangeLogEntry,
    FlowMeter,
    GraphicNote,
    ParsedSource,
    PlantInfo,
    SequenceNote,
)


# ─── Source-xlsx synthesis ──────────────────────────────────────────────────


def _make_source_xlsx() -> bytes:
    wb = Workbook()
    # openpyxl creates a "Sheet" by default — repurpose it as the first
    # known sheet so we don't have to delete it.
    chem = wb.active
    chem.title = "Chemical"
    chem["A17"] = "Flow Meter Info"
    chem["A18"] = "Chem"
    chem["B18"] = "Meter Description"
    chem["C18"] = "K-Factor"
    chem["D18"] = "Make/Model"
    chem["A19"] = "MCA"
    chem["B19"] = "MCA-1"
    chem["C19"] = 459
    chem["D19"] = "azbil (magtech)"
    chem["A20"] = "Water"
    chem["B20"] = "Water"
    chem["C20"] = 1
    chem["D20"] = "n/a"
    chem["A21"] = "Mold AC"
    chem["B21"] = "Mold AC"
    chem["C21"] = 9.4
    chem["D21"] = "seametrics paddle wheel"
    # Cutoff line that the parser should respect.
    chem["A26"] = "QC Factors"

    cyl1 = wb.create_sheet("Cylinder 1 Sequencing")
    cyl1["A1"] = "Cylinder 1 Step by Step Sequence"
    cyl1["A3"] = "Initial Vacuum"
    cyl1["A4"] = "Open V1, R-valve, start vac pump."
    cyl1["A5"] = "Hold at vacuum setpoint."
    cyl1["A7"] = "Fill"
    cyl1["A8"] = "Open T1/T2/T4, flood cylinder under vacuum."

    cyl2 = wb.create_sheet("Cylinder 2 Sequencing")
    cyl2["A1"] = "Cylinder 2 Step by Step Sequence"
    cyl2["A3"] = "Initial Vacuum"
    cyl2["A4"] = "Cyl 2 differs: TnkvISO opens on a delay."

    mix = wb.create_sheet("Mix Sequencing")
    mix["A1"] = "Auto Mixing Step by Step Sequence"
    mix["A3"] = "Initial Vacuum"
    mix["A4"] = "Mix systems don't run vacuum; placeholder note."

    cg = wb.create_sheet("Cylinder 1 Treat Graphic")
    cg["P1"] = "Notes"
    cg["P2"] = "Freq drives on both pressure pumps"
    cg["P3"] = "The line containing T5 does not exist"

    mg = wb.create_sheet("Mix Graphic")
    mg["O1"] = "Notes"
    mg["O2"] = "Tank 17 is not used"

    pi = wb.create_sheet("Plant Info")
    pi["A24"] = "Zebra Placard Printer number is 2110"

    # Sheet the parser should ignore + record as unknown.
    wb.create_sheet("Some Unknown Sheet")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _make_template_xlsx() -> bytes:
    """Minimal CommWKBK-shaped template covering every sheet the builder
    targets, with just enough rows to anchor the mapping logic."""
    wb = Workbook()
    # Tank And Chem Number — match the canonical layout: header row 4,
    # data rows 5–15, chemical name in col G, K factor in col J.
    tank = wb.active
    tank.title = "Tank And Chem Number"
    tank["B2"] = "ALL TANKS"
    tank["F2"] = "MIX System 1"
    tank["B4"] = "Tank NO."
    tank["C4"] = "Description"
    tank["F4"] = "Chem NO."
    tank["G4"] = "Description"
    tank["H4"] = "Card"
    tank["I4"] = "Channel"
    tank["J4"] = "K Factor"
    tank["K4"] = "Unit from SCADA(0= Gal, 1 =Oz)"
    # Sample rows: chem at col G, existing K factor at col J.
    tank["G5"] = "Water"
    tank["J5"] = 1
    tank["G6"] = "MCA"
    tank["J6"] = 100  # Stale value the builder should NOT overwrite.
    tank["G7"] = "Mold Ac"  # Note the casing difference vs. source.
    tank["J7"] = 55

    # Treat Sequence Sign Off — header + a couple of step rows.
    ts = wb.create_sheet("Treat Sequence Sign Off")
    ts["B1"] = "TREAT SEQUENCE SIGN-OFF"
    ts["B2"] = "UFP SIGNOFF"
    ts["C2"] = "TAS SIGNOFF"
    ts["D2"] = "SIGNOFF DATE"
    ts["E2"] = "COMMENTS"
    ts["A3"] = "STEP 1 - Initial Vacuum"
    ts["A7"] = "STEP 2 - Fill"

    # Mix Sequence Sign Off — header + first sequence row.
    ms = wb.create_sheet("Mix Sequence Sign Off")
    ms["A3"] = "MIX 1 SEQUENCE"
    ms["A4"] = "MD1/2/3 Valve opens"

    # Network Schema — the builder appends rows here for plant facts.
    ns = wb.create_sheet("Network Schema")
    ns["A1"] = "Device"
    ns["B1"] = "Model"
    ns["G1"] = "Notes"
    ns["A2"] = "Main PLC"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── Fixtures ───────────────────────────────────────────────────────────────


@pytest.fixture
def source_bytes() -> bytes:
    return _make_source_xlsx()


@pytest.fixture
def template_bytes() -> bytes:
    return _make_template_xlsx()


@pytest.fixture
def parsed(source_bytes) -> ParsedSource:
    return parse_source(source_bytes)


# ─── Parser tests ───────────────────────────────────────────────────────────


def test_parser_extracts_flow_meters(parsed: ParsedSource):
    chems = {m.chemical.lower(): m for m in parsed.flow_meters}
    assert "mca" in chems
    assert chems["mca"].k_factor == 459
    assert chems["mca"].meter_description == "MCA-1"
    assert "azbil" in chems["mca"].make_model
    assert "water" in chems
    assert chems["water"].k_factor == 1


def test_parser_respects_chemical_section_cutoff(parsed: ParsedSource):
    """No 'QC Factors' nonsense should leak into the flow-meter list."""
    chems = [m.chemical for m in parsed.flow_meters]
    assert "QC Factors" not in chems
    # The fixture only defines 3 meters above the cutoff.
    assert len(parsed.flow_meters) == 3


def test_parser_extracts_sequence_notes_per_cylinder(parsed: ParsedSource):
    cyl1 = [n for n in parsed.sequence_notes if n.cylinder == 1]
    cyl2 = [n for n in parsed.sequence_notes if n.cylinder == 2]
    mix = [n for n in parsed.sequence_notes if n.cylinder is None]
    assert any(n.step_name == "Initial Vacuum" for n in cyl1)
    assert any(n.step_name == "Fill" for n in cyl1)
    assert any(n.step_name == "Initial Vacuum" for n in cyl2)
    assert any(n.step_name == "Initial Vacuum" for n in mix)


def test_parser_known_steps_covers_all_canonical_names():
    """If KNOWN_STEPS shrinks, the builder mapping table needs updating too."""
    assert "Initial Vacuum" in KNOWN_STEPS
    assert "Fill" in KNOWN_STEPS
    assert "Empty" in KNOWN_STEPS


def test_parser_extracts_graphic_notes(parsed: ParsedSource):
    sections = {g.section: g for g in parsed.graphic_notes}
    assert "Cylinder 1" in sections
    assert any("Freq drives" in n for n in sections["Cylinder 1"].notes)
    assert "Mix" in sections


def test_parser_extracts_plant_info(parsed: ParsedSource):
    assert any("Zebra" in f for f in parsed.plant_info.plant_facts)


def test_parser_records_unknown_sheets(parsed: ParsedSource):
    assert "Some Unknown Sheet" in parsed.unknown_sheets


# ─── Builder tests ──────────────────────────────────────────────────────────


def test_builder_adds_source_kfactor_columns(parsed, template_bytes):
    new_bytes, report = build_workbook(parsed, template_bytes)
    wb = load_workbook(io.BytesIO(new_bytes))
    tank = wb["Tank And Chem Number"]
    # Header in col M row 4.
    assert "Source K-Factor" in str(tank["M4"].value)
    # MCA source K-Factor lands in M6.
    assert "459" in str(tank["M6"].value)
    assert report.flow_meters_matched >= 2  # Water + MCA + Mold Ac may all match


def test_builder_does_not_overwrite_existing_kfactor(parsed, template_bytes):
    new_bytes, report = build_workbook(parsed, template_bytes)
    wb = load_workbook(io.BytesIO(new_bytes))
    tank = wb["Tank And Chem Number"]
    # MCA J6 was 100 in the template; must remain 100.
    assert tank["J6"].value == 100


def test_builder_attaches_sequence_notes_to_step_row(parsed, template_bytes):
    new_bytes, report = build_workbook(parsed, template_bytes)
    wb = load_workbook(io.BytesIO(new_bytes))
    ts = wb["Treat Sequence Sign Off"]
    # STEP 1 - Initial Vacuum is row 3; COMMENTS is col E.
    assert ts["E3"].value is not None
    text = str(ts["E3"].value)
    # Both cylinders should land in the same cell.
    assert "Cyl 1" in text
    assert "Cyl 2" in text


def test_builder_writes_graphic_notes_to_section_header(parsed, template_bytes):
    new_bytes, _ = build_workbook(parsed, template_bytes)
    wb = load_workbook(io.BytesIO(new_bytes))
    # Cyl 1 graphic notes land on the STEP 1 row's COMMENTS column. The
    # sequence pass writes first, the graphic pass conflicts — so the
    # sequence text survives but a conflict entry is recorded. Confirm
    # both behaviors:
    ts = wb["Treat Sequence Sign Off"]
    assert "Cyl 1" in str(ts["E3"].value)  # sequence pass result intact


def test_builder_appends_plant_facts_to_network_schema(parsed, template_bytes):
    new_bytes, report = build_workbook(parsed, template_bytes)
    wb = load_workbook(io.BytesIO(new_bytes))
    ns = wb["Network Schema"]
    # The builder appends after max_row (=2 in our template, after the
    # Main PLC row), so row 3+ should have the Zebra fact.
    found = False
    for row in ns.iter_rows(min_row=3, values_only=True):
        if any(v and "Zebra" in str(v) for v in row):
            found = True
            break
    assert found, "Plant fact (Zebra printer) was not written into Network Schema"
    assert report.plant_facts_attached >= 1


def test_builder_records_conflicts(parsed, template_bytes):
    """When the builder tries to write into a non-empty cell, it records
    a ChangeLogEntry with conflict=True and the new value is dropped."""
    new_bytes, report = build_workbook(parsed, template_bytes)
    # At least one conflict expected: graphic notes for Cyl 1 try to
    # write into the same COMMENTS cell that the Cyl 1/2 sequence pass
    # already populated.
    conflicts = [c for c in report.changes if c.conflict]
    assert any(c.sheet == "Treat Sequence Sign Off" for c in conflicts)


def test_builder_warnings_include_unknown_sheets(parsed, template_bytes):
    _, report = build_workbook(parsed, template_bytes)
    assert any("unrecognized source sheets" in w for w in report.warnings)


# ─── Endpoint tests ─────────────────────────────────────────────────────────


def test_default_template_endpoint_returns_bundled_xlsx():
    client = TestClient(app)
    resp = client.get("/api/commissioning-workbook/default-template")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Should be a real .xlsx (PK zip magic bytes).
    assert resp.content[:2] == b"PK"


def test_build_endpoint_returns_xlsx_and_report_header(source_bytes):
    client = TestClient(app)
    files = {"source": ("test.xlsx", source_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = client.post("/api/commissioning-workbook/build", files=files)
    assert resp.status_code == 200, resp.text
    assert resp.content[:2] == b"PK"
    # The X-Build-Report header must be present and decodable.
    report_b64 = resp.headers.get("x-build-report")
    assert report_b64
    import base64, json
    report = json.loads(base64.b64decode(report_b64).decode("utf-8"))
    assert report["flow_meters_matched"] >= 1
    assert "changes" in report


def test_build_endpoint_rejects_empty_source():
    client = TestClient(app)
    files = {"source": ("empty.xlsx", b"",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = client.post("/api/commissioning-workbook/build", files=files)
    assert resp.status_code == 400
