"""Tests for the Ignition Tag Builder (backend/features/ignition_tags)."""

from __future__ import annotations

import io
import json
from pathlib import Path

from openpyxl import Workbook

from backend.features.ignition_tags.builder import (
	build_all,
	build_nested_structure,
)
from backend.features.ignition_tags.packager import build_ignition_tree, sort_tree
from backend.features.ignition_tags.parser import parse_workbook

FIXTURES = Path(__file__).parent / "fixtures" / "ignition_tags"


# ---------------------------------------------------------------------------
# Helpers — build small ad-hoc workbooks in memory for the unit tests
# ---------------------------------------------------------------------------

def _make_workbook(
	provider: str | None,
	site: str | None,
	sheets: list[dict],
) -> bytes:
	"""Build an xlsx in memory for tests.

	`sheets` is a list of dicts with keys: name, udt_type, folder,
	headers (list of strings), rows (list of lists, len matching headers).
	A None for udt_type or folder leaves the corresponding cell blank.

	Provider/site/udt_type/folder go in column C (the Jython reads POI
	col=2 → column C). The B-column labels are optional and skipped
	here for brevity.
	"""
	wb = Workbook()
	header = wb.active
	header.title = "Config"
	if provider is not None:
		header["C2"] = provider
	if site is not None:
		header["C3"] = site

	for sheet_spec in sheets:
		ws = wb.create_sheet(title=sheet_spec["name"])
		if sheet_spec.get("udt_type") is not None:
			ws["C2"] = sheet_spec["udt_type"]
		if sheet_spec.get("folder") is not None:
			ws["C3"] = sheet_spec["folder"]
		for i, h in enumerate(sheet_spec["headers"]):
			ws.cell(row=1, column=5 + i, value=h)
		for r, row in enumerate(sheet_spec["rows"]):
			for c, v in enumerate(row):
				if v is None:
					continue
				ws.cell(row=2 + r, column=5 + c, value=v)
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


# ---------------------------------------------------------------------------
# Golden test — the contract test
# ---------------------------------------------------------------------------

USER_GOLDEN_INPUT = FIXTURES / "golden_input.xlsx"
USER_GOLDEN_EXPECTED = FIXTURES / "golden_expected_tree.json"


def test_golden_tree_matches_expected():
	"""Contract test: the rooted nested folder tree the builder produces
	must match what Ignition Designer's Tag Browser Import expects."""
	file_bytes = USER_GOLDEN_INPUT.read_bytes()
	parsed = parse_workbook(file_bytes)
	instances, report = build_all(parsed)
	assert not report.has_errors, [e.model_dump() for e in report.errors]

	actual = sort_tree(build_ignition_tree(instances))
	expected = json.loads(USER_GOLDEN_EXPECTED.read_text())
	assert actual == expected


# ---------------------------------------------------------------------------
# Unit tests — §7.5
# ---------------------------------------------------------------------------

def test_empty_cell_skips_tag():
	"""Tank 3 in the golden fixture has blank Raw Max and Scaled Offset."""
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _report = build_all(parsed)
	tank3 = next(
		i for i in instances
		if i.instance.name == "Tank 3"
	)
	tag_names = {t.name for t in tank3.instance.tags}
	assert tag_names == {"Raw Min", "Scaled Min", "Scaled Max"}
	assert "Raw Max" not in tag_names
	assert "Scaled Offset" not in tag_names


def test_dot_notation_creates_nested_folders():
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Status.Running"],
				"rows": [["T1", "A1", "01", "STATUS"]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	instances, report = build_all(parsed)
	assert not report.has_errors
	tags = instances[0].instance.tags
	assert len(tags) == 1
	folder = tags[0]
	assert folder.name == "Status"
	assert folder.tagType == "Folder"
	assert len(folder.tags) == 1
	atomic = folder.tags[0]
	assert atomic.name == "Running"
	assert atomic.opcItemPath.binding == "ns=1;s=[{plc}]STATUS"


def test_deeply_nested_dot_notation():
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": [
					"Name", "System Name", "System Number", "Setpoints.High.HH"
				],
				"rows": [["T1", "A1", "01", "HH_TAG"]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	instances, report = build_all(parsed)
	assert not report.has_errors

	tags = instances[0].instance.tags
	setpoints = tags[0]
	assert setpoints.name == "Setpoints"
	high = setpoints.tags[0]
	assert high.name == "High"
	hh = high.tags[0]
	assert hh.name == "HH"
	assert hh.opcItemPath.binding == "ns=1;s=[{plc}]HH_TAG"


def test_sibling_folders_merge():
	"""Two headers under the same parent share one folder, not two."""
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": [
					"Name", "System Name", "System Number",
					"Status.Running", "Status.Fault",
				],
				"rows": [["T1", "A1", "01", "R", "F"]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	instances, report = build_all(parsed)
	assert not report.has_errors
	tags = instances[0].instance.tags
	# Exactly one Status folder, containing both children.
	assert len(tags) == 1
	folder = tags[0]
	assert folder.name == "Status"
	child_names = sorted(t.name for t in folder.tags)
	assert child_names == ["Fault", "Running"]


def test_numeric_cell_integer_coercion():
	"""550.0 in the cell becomes '550' in the binding, not '550.0'."""
	# openpyxl normally stores a literal `550` as int, but Excel often
	# serializes integers as floats. Exercise the float path explicitly.
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _ = build_all(parsed)
	tank1 = next(i for i in instances if i.instance.name == "Tank 1")
	tag_by_name = {t.name: t for t in tank1.instance.tags}
	assert tag_by_name["Raw Max"].opcItemPath.binding == "ns=1;s=[{plc}]550"

	# And a direct unit test on build_nested_structure with a float
	tags: list[dict] = []
	build_nested_structure(tags, [], "T", 550.0)
	assert tags[0]["opcItemPath"]["binding"] == "ns=1;s=[{plc}]550"


def test_required_columns_in_any_order_is_accepted():
	"""Real production workbooks put `System Name` / `System Number` /
	`Name` in that order — the Jython does name-based lookup and
	doesn't care about order. The port must accept the same."""
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Pump",
				"udt_type": "Pumps/Pump",
				"folder": "Edge/Pumps",
				"headers": [
					"System Name", "System Number", "Name",  # reordered
					"Manual", "Outputs",
				],
				"rows": [
					["Mixing", 2, "ECO", "MW3053.10", "MW3054.10"],
				],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	instances, report = build_all(parsed)
	assert not report.has_errors, [e.model_dump() for e in report.errors]
	assert len(instances) == 1
	inst = instances[0]
	assert inst.base_path == "[Athens]UFP_Athens/Mixing/2/Edge/Pumps"
	assert inst.instance.name == "ECO"
	tag_names = {t.name for t in inst.instance.tags}
	assert tag_names == {"Manual", "Outputs"}


def test_required_columns_missing_is_error():
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "Raw Min"],  # missing System Number
				"rows": [["T1", "A1", "01"]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	_instances, report = build_all(parsed)
	assert report.has_errors
	codes = [e.code for e in report.errors]
	assert "sheet.missing_required_column" in codes


def test_blank_c2_c3_on_data_sheet_aborts_cleanly():
	"""C2 is the UDT type id, C3 is the destination folder (the Jython
	reads POI col=2, which is spreadsheet column C). Blank → fatal."""
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": None,  # blank C2
				"folder": None,    # blank C3
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [["T1", "A1", "01", 500]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	_instances, report = build_all(parsed)
	codes = {e.code for e in report.errors}
	assert "sheet.missing_udt_type" in codes
	assert "sheet.missing_folder" in codes


def test_duplicate_instance_warning():
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [
					["Tank A", "A1", "01", 500],
					["Tank A", "A1", "01", 501],  # duplicate path+name
				],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	_instances, report = build_all(parsed)
	assert not report.has_errors
	codes = [w.code for w in report.warnings]
	assert "duplicate_instance" in codes


def test_workbook_with_only_header_sheet_is_error():
	wb = Workbook()
	wb.active.title = "Header"
	wb.active["B2"] = "Athens"
	wb.active["B3"] = "UFP_Athens"
	buf = io.BytesIO()
	wb.save(buf)
	parsed = parse_workbook(buf.getvalue())
	codes = {e.code for e in parsed.issues}
	assert "workbook.too_few_sheets" in codes


def test_header_sheet_missing_provider_or_site_is_error():
	xlsx = _make_workbook(
		None,  # no provider in B2
		None,  # no site in B3
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [["T1", "A1", "01", 500]],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	codes = {e.code for e in parsed.issues}
	assert "header.missing_provider" in codes
	assert "header.missing_site" in codes


def test_no_data_rows_is_warning_not_error():
	"""An empty data sheet warns but does not abort the workbook.

	The Jython silently iterates zero rows for an empty sheet; the
	port matches that, demoting "no data rows" from error to warning.
	"""
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [],  # nothing
			}
		],
	)
	parsed = parse_workbook(xlsx)
	codes = {i.code for i in parsed.issues}
	assert "sheet.no_data_rows" in codes
	# No errors — only the warning.
	errors = [i for i in parsed.issues if i.severity == "error"]
	assert errors == []


def test_table_terminates_at_blank_name():
	"""A blank Name cell ends the data table — later rows are not parsed."""
	xlsx = _make_workbook(
		"Athens",
		"UFP_Athens",
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [
					["T1", "A1", "01", 500],
					[None, "A1", "01", 501],  # blank Name -> ends here
					["T3", "A1", "01", 502],  # should be ignored
				],
			}
		],
	)
	parsed = parse_workbook(xlsx)
	instances, report = build_all(parsed)
	assert not report.has_errors
	names = sorted(i.instance.name for i in instances)
	assert names == ["T1"]


def test_base_path_format_matches_jython():
	"""[<provider>]<site>/<sys_name>/<sys_num>/<folder>"""
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _ = build_all(parsed)
	tank1 = next(i for i in instances if i.instance.name == "Tank 1")
	assert tank1.base_path == "[Athens]UFP_Athens/A1/01/LevelSensors"


def test_atomic_tag_keys_match_ignition_import_format():
	"""extra='forbid' guards against accidental key additions.

	Every atomic tag carries exactly `name`, `tagType`, `opcItemPath`
	— the shape Ignition Designer's Tag Browser Import expects.
	"""
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _ = build_all(parsed)
	tank1 = next(i for i in instances if i.instance.name == "Tank 1")
	for tag in tank1.instance.tags:
		dumped = tag.model_dump()
		assert set(dumped.keys()) == {"name", "tagType", "opcItemPath"}
		assert dumped["tagType"] == "AtomicTag"


# ---------------------------------------------------------------------------
# Packager + e2e
# ---------------------------------------------------------------------------

def test_build_ignition_tree_nests_by_path_segments():
	"""The flat (base_path, instance) list collapses into a rooted
	folder tree: site → sys_name → sys_num → folder → instance."""
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _report = build_all(parsed)
	tree = build_ignition_tree(instances)

	# Root is the site name (provider prefix is stripped).
	assert tree["name"] == "UFP_Athens"
	assert tree["tagType"] == "Folder"

	# Walk down to the Tank Level Sensors folder and assert the three
	# Tank instances land there.
	def find_child(folder, name):
		return next(c for c in folder["tags"] if c["name"] == name)

	a1 = find_child(tree, "A1")
	one = find_child(a1, "01")
	level_sensors = find_child(one, "LevelSensors")
	tank_names = {c["name"] for c in level_sensors["tags"] if c.get("tagType") == "UdtInstance"}
	assert tank_names == {"Tank 1", "Tank 2", "Tank 3"}


def test_tree_drops_provider_prefix():
	"""Provider (e.g. `[SCADA]`) is chosen at import time in Designer
	and must not appear in the tree."""
	parsed = parse_workbook(USER_GOLDEN_INPUT.read_bytes())
	instances, _ = build_all(parsed)
	tree = build_ignition_tree(instances)
	# Recursively collect every folder/instance name.
	names: list[str] = []
	def walk(n):
		names.append(n["name"])
		for c in n.get("tags", []):
			walk(c)
	walk(tree)
	assert not any("[" in n or "]" in n for n in names), (
		"Provider prefix must not leak into folder names"
	)


def test_e2e_router_returns_json_with_bundle_and_report():
	"""End-to-end: POST the synthetic fixture, get a JSON envelope with
	`bundle` + `validation_report` + metadata."""
	from fastapi.testclient import TestClient
	from backend.api.main import app

	client = TestClient(app)
	with USER_GOLDEN_INPUT.open("rb") as f:
		response = client.post(
			"/api/ignition-tags/build",
			files={
				"file": (
					"golden_input.xlsx",
					f,
					"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
				)
			},
		)
	assert response.status_code == 200, response.text
	assert response.headers["content-type"].startswith("application/json")

	body = response.json()
	assert set(body.keys()) >= {"bundle", "validation_report", "site", "instance_count"}
	assert body["site"] == "UFP_Athens"
	# Bundle is a single rooted folder tree, not a flat map.
	assert body["bundle"]["name"] == "UFP_Athens"
	assert body["bundle"]["tagType"] == "Folder"
	# instance_count is the number of UdtInstance leaves in the tree.
	def count_instances(node):
		if node.get("tagType") == "UdtInstance":
			return 1
		return sum(count_instances(c) for c in node.get("tags", []))
	assert count_instances(body["bundle"]) == body["instance_count"]
	assert body["instance_count"] > 0
	assert body["validation_report"]["errors"] == []


def test_e2e_router_returns_400_on_bad_workbook():
	"""A workbook with structural errors yields 400 + a validation report."""
	from fastapi.testclient import TestClient
	from backend.api.main import app

	client = TestClient(app)
	bad = _make_workbook(
		None,  # blank provider — fatal
		None,  # blank site — fatal
		[
			{
				"name": "Sheet1",
				"udt_type": "Tank/Tank Level Sensors",
				"folder": "F",
				"headers": ["Name", "System Name", "System Number", "Raw Min"],
				"rows": [["T1", "A1", "01", 500]],
			}
		],
	)
	response = client.post(
		"/api/ignition-tags/build",
		files={
			"file": (
				"bad.xlsx",
				bad,
				"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			)
		},
	)
	assert response.status_code == 400
	body = response.json()
	assert "validation_report" in body
	codes = {e["code"] for e in body["validation_report"]["errors"]}
	assert "header.missing_provider" in codes
	assert "header.missing_site" in codes
